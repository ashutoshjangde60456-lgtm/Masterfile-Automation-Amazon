import io
import json
import re
import time
import zipfile
import xml.etree.ElementTree as ET
from textwrap import dedent
from pathlib import Path

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from difflib import SequenceMatcher

# ─────────────────────────────────────────────────────────────────────
# Page meta + theming
# ─────────────────────────────────────────────────────────────────────
st.set_page_config(page_title="Masterfile Automation - Amazon", page_icon="🧾", layout="wide")
st.markdown("""
<style>
:root{ --bg1:#f6f9fc; --bg2:#fff; --card:#fff; --card-border:#e8eef6;
--ink:#0f172a; --muted:#64748b; --accent:#2563eb; }
.stApp{background:linear-gradient(180deg, var(--bg1) 0%, var(--bg2) 70%);}
.block-container{padding-top:.75rem;}
.section{border:1px solid var(--card-border);background:var(--card);border-radius:16px;
  padding:18px 20px; box-shadow:0 6px 24px rgba(2,6,23,.05); margin-bottom:18px;}
.badge{display:inline-block;padding:4px 10px;border-radius:999px;font-size:.82rem;font-weight:600;margin-right:.25rem}
.badge-info{background:#eef2ff;color:#1e40af} .badge-ok{background:#ecfdf5;color:#065f46}
div.stButton>button,.stDownloadButton>button{background:var(--accent)!important;color:#fff!important;border-radius:10px!important;border:0!important}
</style>
""", unsafe_allow_html=True)

# ─────────────────────────────────────────────────────────────────────
# Template layout constants
# ─────────────────────────────────────────────────────────────────────
MASTER_TEMPLATE_SHEET = "Template"   # target sheet
MASTER_DISPLAY_ROW    = 2            # human headers
MASTER_SECONDARY_ROW  = 3            # bullet disambiguators
MASTER_DATA_START_ROW = 4            # first data row

# ─────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────
XL_NS_MAIN = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
XL_NS_REL  = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
ET.register_namespace("", XL_NS_MAIN)
ET.register_namespace("r", XL_NS_REL)
ET.register_namespace("mc", "http://schemas.openxmlformats.org/markup-compatibility/2006")
ET.register_namespace("x14ac", "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac")

_INVALID_XML_CHARS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F\uD800-\uDFFF]")

def sanitize_xml_text(s) -> str:
    if s is None: return ""
    return _INVALID_XML_CHARS.sub("", str(s))

def norm(s: str) -> str:
    if s is None: return ""
    x = str(s).strip().lower()
    x = re.sub(r"\s*-\s*en\s*[-_ ]\s*us\s*$", "", x)
    x = x.replace("–","-").replace("—","-").replace("−","-")
    x = re.sub(r"[._/\\-]+", " ", x)
    x = re.sub(r"[^0-9a-z\s]+", " ", x)
    return re.sub(r"\s+", " ", x).strip()

def top_matches(query, candidates, k=3):
    q = norm(query)
    scored = [(SequenceMatcher(None, q, norm(c)).ratio(), c) for c in candidates]
    scored.sort(key=lambda t: t[0], reverse=True)
    return scored[:k]

def nonempty_rows(df: pd.DataFrame) -> int:
    if df.empty: return 0
    return df.replace("", pd.NA).dropna(how="all").shape[0]

def worksheet_used_cols(ws, header_rows=(1,), hard_cap=2048, empty_streak_stop=8):
    max_try = min(ws.max_column, hard_cap)
    last_nonempty, streak = 0, 0
    for c in range(1, max_try + 1):
        any_val = any((ws.cell(row=r, column=c).value not in (None, "")) for r in header_rows)
        if any_val: last_nonempty, streak = c, 0
        else:
            streak += 1
            if streak >= empty_streak_stop: break
    return max(last_nonempty, 1)

def _col_letter(n: int) -> str:
    s = ""
    while n:
        n, r = divmod(n-1, 26)
        s = chr(65+r) + s
    return s

def _col_number(letters: str) -> int:
    n = 0
    for ch in letters:
        if not ch.isalpha(): break
        n = n * 26 + (ord(ch.upper()) - 64)
    return n

# ── ZIP / XML helpers ────────────────────────────────────────────────
def _find_sheet_part_path(z: zipfile.ZipFile, sheet_name: str) -> str:
    wb_xml = ET.fromstring(z.read("xl/workbook.xml"))
    rels_xml = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
    rid = None
    for sh in wb_xml.find(f"{{{XL_NS_MAIN}}}sheets"):
        if sh.attrib.get("name") == sheet_name:
            rid = sh.attrib.get(f"{{{XL_NS_REL}}}id")
            break
    if not rid: raise ValueError(f"Sheet '{sheet_name}' not found.")
    target = None
    for rel in rels_xml:
        if rel.attrib.get("Id") == rid:
            target = rel.attrib.get("Target")
            break
    if not target: raise ValueError(f"Relationship for sheet '{sheet_name}' not found.")
    target = target.replace("\\", "/")
    if target.startswith("../"): target = target[3:]
    if not target.startswith("xl/"): target = "xl/" + target
    return target  # e.g., xl/worksheets/sheet1.xml

def _get_table_paths_for_sheet(z: zipfile.ZipFile, sheet_path: str) -> list:
    rels_path = sheet_path.replace("worksheets/", "worksheets/_rels/").replace(".xml", ".xml.rels")
    if rels_path not in z.namelist(): return []
    root = ET.fromstring(z.read(rels_path))
    out = []
    for rel in root:
        t = rel.attrib.get("Type", "")
        if t.endswith("/table"):
            target = rel.attrib.get("Target", "").replace("\\", "/")
            if target.startswith("../"): target = target[3:]
            if not target.startswith("xl/"): target = "xl/" + target
            out.append(target)
    return out

def _read_table_cols_count(table_xml_bytes: bytes) -> int:
    try:
        root = ET.fromstring(table_xml_bytes)
        tcols = root.find(f"{{{XL_NS_MAIN}}}tableColumns")
        if tcols is None: return 0
        cnt_attr = tcols.attrib.get("count")
        cnt = int(cnt_attr) if cnt_attr else 0
        child_count = sum(1 for _ in tcols)
        return max(cnt, child_count)
    except Exception:
        return 0

def _union_dimension(orig_dim_ref: str, used_cols: int, last_row: int) -> str:
    try:
        _, right = orig_dim_ref.split(":", 1)
        m = re.match(r"([A-Z]+)(\d+)", right)
        if m:
            orig_last_col = _col_number(m.group(1))
            orig_last_row = int(m.group(2))
        else:
            orig_last_col, orig_last_row = used_cols, last_row
    except Exception:
        orig_last_col, orig_last_row = used_cols, last_row
    u_last_col = max(orig_last_col, used_cols)
    u_last_row = max(orig_last_row, last_row)
    return f"A1:{_col_letter(u_last_col)}{u_last_row}"

def _ensure_ws_x14ac(root):
    # Allow x14ac attributes without repairs
    root.set("{http://schemas.openxmlformats.org/markup-compatibility/2006}Ignorable", "x14ac")

def _intersects_range(a1: str, r1: int, r2: int) -> bool:
    # a1 like "A3:B7" → True if overlap with [r1, r2]
    m = re.match(r"^[A-Z]+(\d+):[A-Z]+(\d+)$", a1 or "", re.I)
    if not m:
        return False
    lo = int(m.group(1)); hi = int(m.group(2))
    if lo > hi: lo, hi = hi, lo
    return not (hi < r1 or lo > r2)

def _patch_sheet_xml(sheet_xml_bytes: bytes, header_row: int, start_row: int, used_cols_final: int, block_2d: list) -> bytes:
    root = ET.fromstring(sheet_xml_bytes)
    _ensure_ws_x14ac(root)

    sheetData = root.find(f"{{{XL_NS_MAIN}}}sheetData")
    if sheetData is None:
        sheetData = ET.SubElement(root, f"{{{XL_NS_MAIN}}}sheetData")

    # 1) Remove existing data rows at/after start_row
    for row in list(sheetData):
        try:
            r = int(row.attrib.get("r") or "0")
        except Exception:
            r = 0
        if r >= start_row:
            sheetData.remove(row)

    # 2) Remove mergeCells that intersect our data region to prevent "Repaired Records"
    mergeCells = root.find(f"{{{XL_NS_MAIN}}}mergeCells")
    if mergeCells is not None:
        for mc in list(mergeCells):
            ref = mc.attrib.get("ref", "")
            if _intersects_range(ref, start_row, 1048576):
                mergeCells.remove(mc)
        if len(list(mergeCells)) == 0:
            root.remove(mergeCells)

    # 3) Append dense rows (A..lastCol) using inlineStr (keeps rows visible, no sparse-row repair)
    row_span = f"1:{used_cols_final}" if used_cols_final > 0 else "1:1"
    n_rows = len(block_2d)
    for i in range(n_rows):
        r = start_row + i
        src_row = block_2d[i]
        row_el = ET.Element(f"{{{XL_NS_MAIN}}}row", r=str(r))
        row_el.set("spans", row_span)
        row_el.set("{http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac}dyDescent", "0.25")

        for j in range(used_cols_final):
            val = src_row[j] if j < len(src_row) else ""
            txt = sanitize_xml_text(val) if val else ""
            col = _col_letter(j + 1)
            c = ET.Element(f"{{{XL_NS_MAIN}}}c", r=f"{col}{r}", t="inlineStr")
            is_el = ET.SubElement(c, f"{{{XL_NS_MAIN}}}is")
            t_el = ET.SubElement(is_el, f"{{{XL_NS_MAIN}}}t")
            t_el.set("{http://www.w3.org/XML/1998/namespace}space", "preserve")
            t_el.text = txt  # empty allowed → row still visible
            row_el.append(c)

        sheetData.append(row_el)

    # 4) Dimension: conservative union with original
    dim = root.find(f"{{{XL_NS_MAIN}}}dimension")
    if dim is None:
        dim = ET.SubElement(root, f"{{{XL_NS_MAIN}}}dimension", ref="A1:A1")
    last_row = max(header_row, start_row + max(0, n_rows) - 1)
    new_ref = _union_dimension(dim.attrib.get("ref", "A1:A1"), used_cols_final, last_row)
    dim.set("ref", new_ref)

    # 5) AutoFilter: only update if one existed originally
    af = root.find(f"{{{XL_NS_MAIN}}}autoFilter")
    if af is not None:
        af.set("ref", f"A{header_row}:{_col_letter(used_cols_final)}{last_row}")

    # 6) Clear filterMode flag if present (prevents repair on changed rows)
    sheetPr = root.find(f"{{{XL_NS_MAIN}}}sheetPr")
    if sheetPr is not None and sheetPr.attrib.get("filterMode"):
        sheetPr.attrib.pop("filterMode", None)

    return ET.tostring(root, encoding="utf-8", xml_declaration=True)

def _patch_table_xml(table_xml_bytes: bytes, header_row: int, last_row: int, last_col_n: int) -> bytes:
    root = ET.fromstring(table_xml_bytes)
    new_ref = f"A{header_row}:{_col_letter(last_col_n)}{last_row}"
    root.set("ref", new_ref)

    af = root.find(f"{{{XL_NS_MAIN}}}autoFilter")
    if af is None:
        af = ET.SubElement(root, f"{{{XL_NS_MAIN}}}autoFilter")
    af.set("ref", new_ref)

    # Keep tableColumns list as-is; just ensure the 'count' equals the number of children (Excel requirement)
    tcols = root.find(f"{{{XL_NS_MAIN}}}tableColumns")
    if tcols is not None:
        child_count = sum(1 for _ in tcols)
        tcols.set("count", str(child_count))
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)

def _strip_calcchain_override(ct_bytes: bytes) -> bytes:
    try:
        ns = "http://schemas.openxmlformats.org/package/2006/content-types"
        root = ET.fromstring(ct_bytes)
        ET.register_namespace("", ns)
        for el in list(root):
            if el.tag == f"{{{ns}}}Override" and el.attrib.get("PartName","").lower() == "/xl/calcchain.xml":
                root.remove(el)
        return ET.tostring(root, encoding="utf-8", xml_declaration=True)
    except Exception:
        return ct_bytes

def fast_patch_template(master_bytes: bytes, sheet_name: str, header_row: int, start_row: int, used_cols: int, block_2d: list) -> bytes:
    """Ultra-fast writer: swaps only the target sheet XML + syncs tables & filters; removes calcChain."""
    zin = zipfile.ZipFile(io.BytesIO(master_bytes), "r")
    sheet_path = _find_sheet_part_path(zin, sheet_name)
    table_paths = _get_table_paths_for_sheet(zin, sheet_path)

    # Use at least the widest table width (some tables define more columns than headers)
    max_cols = used_cols
    for tp in table_paths:
        try:
            cnt = _read_table_cols_count(zin.read(tp))
            if cnt > max_cols: max_cols = cnt
        except Exception:
            pass

    new_sheet_xml = _patch_sheet_xml(zin.read(sheet_path), header_row, start_row, max_cols, block_2d)

    last_row = max(header_row, start_row + max(0, len(block_2d)) - 1)
    patched_tables = {}
    for tp in table_paths:
        try:
            patched_tables[tp] = _patch_table_xml(zin.read(tp), header_row, last_row, max_cols)
        except Exception:
            pass

    out_bio = io.BytesIO()
    with zipfile.ZipFile(out_bio, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            fn = item.filename
            if fn == sheet_path:
                zout.writestr(item, new_sheet_xml)
            elif fn in patched_tables:
                zout.writestr(item, patched_tables[fn])
            elif fn.lower() == "[content_types].xml":
                zout.writestr(item, _strip_calcchain_override(zin.read(fn)))
            elif fn.lower() == "xl/calcchain.xml":
                # Drop calcChain so Excel rebuilds without 'repair'
                continue
            else:
                zout.writestr(item, zin.read(fn))
    zin.close()
    out_bio.seek(0)
    return out_bio.getvalue()

# ─────────────────────────────────────────────────────────────────────
# UI — inputs
# ─────────────────────────────────────────────────────────────────────
st.title("🧾 Masterfile Automation – Amazon")
st.caption("Ultra-fast writer (seconds). Preserves all sheets, styles, formulas, and macros (.xlsm).")

st.markdown("<div class='section'><span class='badge badge-info'>Template-only writer</span> "
            "<span class='badge badge-ok'>Fast XML, no fallbacks</span></div>", unsafe_allow_html=True)

st.markdown("<div class='section'>", unsafe_allow_html=True)
c1, c2 = st.columns([1, 1])
with c1:
    masterfile_file = st.file_uploader("📄 Masterfile Template (.xlsx / .xlsm)", type=["xlsx", "xlsm"])
with c2:
    onboarding_file = st.file_uploader("🧾 Onboarding (.xlsx)", type=["xlsx"])

st.markdown("#### 🔗 Mapping JSON")
tab1, tab2 = st.tabs(["Paste JSON", "Upload JSON"])
mapping_json_text, mapping_json_file = "", None
with tab1:
    mapping_json_text = st.text_area("Paste mapping JSON", height=200,
                                     placeholder='\n{\n  "Partner SKU": ["Seller SKU", "item_sku"]\n}\n')
with tab2:
    mapping_json_file = st.file_uploader("Or upload mapping.json", type=["json"], key="mapping_file")
st.markdown("</div>", unsafe_allow_html=True)

st.divider()
go = st.button("🚀 Generate Final Masterfile", type="primary")

# ─────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────
SENTINEL_LIST = object()

if go:
    st.markdown("<div class='section'>", unsafe_allow_html=True)
    st.markdown("### 📝 Log")
    log = st.empty()
    def slog(msg): log.markdown(msg)

    if not masterfile_file or not onboarding_file:
        st.error("Please upload both **Masterfile Template** and **Onboarding**.")
        st.markdown("</div>", unsafe_allow_html=True)
        st.stop()

    # extension & mime
    ext = (Path(masterfile_file.name).suffix or ".xlsx").lower()
    mime_map = {
        ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        ".xlsm": "application/vnd.ms-excel.sheet.macroEnabled.12",
    }
    out_mime = mime_map.get(ext, mime_map[".xlsx"])

    # Parse mapping JSON
    try:
        mapping_raw = json.loads(mapping_json_text) if mapping_json_text.strip() else json.load(mapping_json_file)
    except Exception as e:
        st.error(f"Mapping JSON parse error: {e}")
        st.markdown("</div>", unsafe_allow_html=True)
        st.stop()
    if not isinstance(mapping_raw, dict):
        st.error("Mapping JSON must be an object: {\"Master header\": [aliases...]}.")
        st.markdown("</div>", unsafe_allow_html=True)
        st.stop()

    # Normalize mapping: { master_norm: [aliases...] }
    mapping_aliases = {}
    for k, v in mapping_raw.items():
        aliases = v[:] if isinstance(v, list) else [v]
        if k not in aliases: aliases.append(k)
        mapping_aliases[norm(k)] = aliases

    # Read template headers fast (read-only)
    masterfile_file.seek(0)
    master_bytes = masterfile_file.read()

    slog("⏳ Reading Template headers…")
    t0 = time.time()
    wb_ro = load_workbook(io.BytesIO(master_bytes), read_only=True, data_only=True, keep_links=True)
    if MASTER_TEMPLATE_SHEET not in wb_ro.sheetnames:
        st.error(f"Sheet **'{MASTER_TEMPLATE_SHEET}'** not found in the masterfile."); st.stop()
    ws_ro = wb_ro[MASTER_TEMPLATE_SHEET]
    used_cols = worksheet_used_cols(ws_ro, header_rows=(MASTER_DISPLAY_ROW, MASTER_SECONDARY_ROW), hard_cap=2048, empty_streak_stop=8)
    display_headers   = [ws_ro.cell(row=MASTER_DISPLAY_ROW,   column=c).value or "" for c in range(1, used_cols+1)]
    secondary_headers = [ws_ro.cell(row=MASTER_SECONDARY_ROW, column=c).value or "" for c in range(1, used_cols+1)]
    wb_ro.close()
    slog(f"✅ Headers loaded (cols={used_cols}) in {time.time()-t0:.2f}s")

    # Pick best onboarding sheet
    try:
        best_xl = pd.ExcelFile(onboarding_file)
        best, best_score, best_info = None, -1, ""
        for sheet in best_xl.sheet_names:
            try:
                df = best_xl.parse(sheet_name=sheet, header=0, dtype=str).fillna("")
                df.columns = [str(c).strip() for c in df.columns]
            except Exception:
                continue
            header_set = {norm(c) for c in df.columns}
            matches = sum(any(norm(a) in header_set for a in aliases)
                          for aliases in mapping_aliases.values())
            rows = nonempty_rows(df)
            score = matches + (0.01 if rows > 0 else 0.0)
            if score > best_score:
                best, best_score = (df, sheet), score
                best_info = f"matched headers: {matches}, non-empty rows: {rows}"
        if best is None:
            raise ValueError("No readable onboarding sheet found.")
        best_df, best_sheet, info = best[0], best[1], best_info
    except Exception as e:
        st.error(f"Onboarding error: {e}"); st.stop()

    on_df = best_df.fillna("")
    on_df.columns = [str(c).strip() for c in on_df.columns]
    on_headers = list(on_df.columns)
    st.success(f"Using onboarding sheet: **{best_sheet}** ({info})")

    # Build mapping master col -> source series
    series_by_alias = {norm(h): on_df[h] for h in on_headers}
    report_lines = ["#### 🔎 Mapping Summary (Template)"]
    BULLET_DISP_N = norm("Key Product Features")
    master_to_source = {}

    for c, (disp, sec) in enumerate(zip(display_headers, secondary_headers), start=1):
        disp_norm = norm(disp); sec_norm = norm(sec)
        if disp_norm == BULLET_DISP_N and sec_norm:
            effective_header = sec; label_for_log = f"{disp} ({sec})"
        else:
            effective_header = disp; label_for_log = disp
        eff_norm = norm(effective_header)
        if not eff_norm: continue
        aliases = mapping_aliases.get(eff_norm, [effective_header])
        resolved = None
        for a in aliases:
            s = series_by_alias.get(norm(a))
            if s is not None:
                resolved = s; break
        if resolved is not None:
            master_to_source[c] = resolved
            report_lines.append(f"- ✅ **{label_for_log}** ← `{a}`")
        else:
            if disp_norm == norm("Listing Action (List or Unlist)"):
                master_to_source[c] = SENTINEL_LIST
                report_lines.append(f"- 🟨 **{label_for_log}** ← (will fill `'List'`)")
            else:
                sugg = top_matches(effective_header, on_headers, 3)
                sug_txt = ", ".join(f"`{name}` ({round(sc*100,1)}%)" for sc, name in sugg) if sugg else "*none*"
                report_lines.append(f"- ❌ **{label_for_log}** ← *no match*. Suggestions: {sug_txt}")
    st.markdown("\n".join(report_lines))

    n_rows = len(on_df)

    # Build sanitized 2-D block (dense writer will emit all columns)
    block = [[""] * used_cols for _ in range(n_rows)]
    for col, src in master_to_source.items():
        if src is SENTINEL_LIST:
            for i in range(n_rows): block[i][col-1] = "List"
        else:
            vals = src.astype(str).tolist()
            m = min(len(vals), n_rows)
            for i in range(m):
                v = sanitize_xml_text(vals[i].strip())
                if v and v.lower() not in ("nan", "none", ""):
                    block[i][col-1] = v

    # FAST XML write (no fallback)
    slog("🚀 Writing via fast XML…")
    t_write = time.time()
    out_bytes = fast_patch_template(
        master_bytes=master_bytes,
        sheet_name=MASTER_TEMPLATE_SHEET,
        header_row=MASTER_DISPLAY_ROW,
        start_row=MASTER_DATA_START_ROW,
        used_cols=used_cols,
        block_2d=block
    )
    slog(f"✅ Done in {time.time()-t_write:.2f}s")

    # Download
    st.download_button(
        "⬇️ Download Final Masterfile",
        data=out_bytes,
        file_name=f"final_masterfile{ext}",
        mime=out_mime,
        key="dl_final_fast",
    )
    st.markdown("</div>", unsafe_allow_html=True)

with st.expander("📘 How to use (step-by-step)", expanded=False):
    st.markdown(dedent(f"""
    **This tool**
    - Writes only into `{MASTER_TEMPLATE_SHEET}` and preserves everything else (including macros).
    - Uses a fast XML sheet swap (seconds) — no slow fallbacks.

    **Run**
    1) Upload the Masterfile (.xlsx/.xlsm) and the Onboarding (.xlsx)
    2) Paste/upload Mapping JSON
    3) Click **Generate**
    """))
